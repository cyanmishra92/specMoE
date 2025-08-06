#!/usr/bin/env python3
"""
Generate CSV/Excel Report from Switch Prefetching Analysis

Creates formatted CSV and Excel files with statistical analysis for easy review.
"""

import pandas as pd
import numpy as np
import json
import pickle
from pathlib import Path
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

class ReportGenerator:
    def __init__(self, results_dir: Path):
        self.results_dir = results_dir
        self.strategies = ["A", "B", "C", "D", "E"]
        self.batch_sizes = [1, 2, 4, 8, 16]
        self.strategy_names = {
            "A": "On-Demand",
            "B": "Oracle",
            "C": "Multi-Look",
            "D": "Top-K",
            "E": "Intelligent"
        }
    
    def load_analysis_results(self) -> dict:
        """Load analysis results"""
        analysis_file = self.results_dir / "analysis.json"
        
        if analysis_file.exists():
            with open(analysis_file) as f:
                return json.load(f)
        else:
            print(f"⚠️  Analysis file not found: {analysis_file}")
            print(f"   Run analyze_results.py first")
            return None
    
    def create_summary_table(self, analysis_data: dict) -> pd.DataFrame:
        """Create main summary table"""
        summary_stats = analysis_data["summary_statistics"]
        df = pd.DataFrame(summary_stats)
        
        # Add strategy names
        df['strategy_name'] = df['strategy'].map(self.strategy_names)
        
        # Round numerical values
        numeric_columns = ['latency_mean', 'latency_std', 'hit_rate_mean', 'hit_rate_std',
                          'memory_usage_mean', 'memory_usage_std', 'prefetch_accuracy_mean',
                          'prefetch_accuracy_std', 'experts_loaded_mean', 'experts_loaded_std']
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].round(3)
        
        # Reorder columns
        column_order = ['strategy', 'strategy_name', 'batch_size', 'num_runs',
                       'latency_mean', 'latency_std', 'hit_rate_mean', 'hit_rate_std',
                       'memory_usage_mean', 'memory_usage_std', 'prefetch_accuracy_mean',
                       'prefetch_accuracy_std', 'experts_loaded_mean', 'experts_loaded_std']
        
        existing_columns = [col for col in column_order if col in df.columns]
        return df[existing_columns]
    
    def create_comparison_matrix(self, analysis_data: dict) -> pd.DataFrame:
        """Create strategy comparison matrix"""
        summary_stats = analysis_data["summary_statistics"]
        df = pd.DataFrame(summary_stats)
        
        # Pivot to create comparison matrix
        pivot_df = df.pivot(index='batch_size', columns='strategy', values='latency_mean')
        
        # Add strategy names to columns
        pivot_df.columns = [f"{col}_{self.strategy_names[col]}" for col in pivot_df.columns]
        
        # Calculate improvements over baseline (Strategy A)
        if 'A_On-Demand' in pivot_df.columns:
            baseline = pivot_df['A_On-Demand']
            for col in pivot_df.columns:
                if col != 'A_On-Demand':
                    improvement_col = f"{col}_Improvement"
                    pivot_df[improvement_col] = ((baseline - pivot_df[col]) / baseline * 100).round(1)
        
        return pivot_df
    
    def create_statistical_tests_table(self, analysis_data: dict) -> pd.DataFrame:
        """Create statistical significance test results table"""
        stat_tests = analysis_data.get("statistical_tests", {})
        
        rows = []
        for batch_key, batch_tests in stat_tests.items():
            batch_size = batch_key.replace("batch_", "")
            
            for comparison, test_result in batch_tests.items():
                strategy1, strategy2 = comparison.split("_vs_")
                
                rows.append({
                    "batch_size": int(batch_size),
                    "strategy_1": strategy1,
                    "strategy_1_name": self.strategy_names[strategy1],
                    "strategy_2": strategy2,
                    "strategy_2_name": self.strategy_names[strategy2],
                    "p_value": round(test_result["p_value"], 6),
                    "significant": test_result["significant"],
                    "effect_size_ms": round(test_result["effect_size"], 3)
                })
        
        return pd.DataFrame(rows)
    
    def create_insights_summary(self, analysis_data: dict) -> pd.DataFrame:
        """Create insights summary table"""
        insights = analysis_data.get("insights", {})
        
        rows = []
        
        # Best overall strategy
        if insights.get("best_overall_strategy"):
            best = insights["best_overall_strategy"]
            rows.append({
                "metric": "Best Overall Strategy",
                "strategy": best["strategy"],
                "strategy_name": self.strategy_names[best["strategy"]],
                "value": f"{best['mean_latency']:.2f}ms",
                "description": "Lowest average latency across all batch sizes"
            })
        
        # Best low latency
        if insights.get("best_low_latency_strategy"):
            best = insights["best_low_latency_strategy"]
            rows.append({
                "metric": "Best Low Latency (Batch=1)",
                "strategy": best["strategy"],
                "strategy_name": self.strategy_names[best["strategy"]],
                "value": f"{best['latency']:.2f} ± {best['std']:.2f}ms",
                "description": "Best performance for single requests"
            })
        
        # Best memory efficient
        if insights.get("best_memory_efficient_strategy"):
            best = insights["best_memory_efficient_strategy"]
            rows.append({
                "metric": "Best Memory Efficiency",
                "strategy": best["strategy"],
                "strategy_name": self.strategy_names[best["strategy"]],
                "value": f"{best['efficiency']:.3f} hit/GB",
                "description": f"Batch size {best['batch_size']}"
            })
        
        # Prefetching effectiveness
        prefetch_eff = insights.get("prefetching_effectiveness", {})
        if prefetch_eff:
            best_prefetch = max(prefetch_eff.items(), key=lambda x: x[1]["average_improvement"])
            rows.append({
                "metric": "Best Prefetching Strategy",
                "strategy": best_prefetch[0],
                "strategy_name": self.strategy_names[best_prefetch[0]],
                "value": f"{best_prefetch[1]['average_improvement']:.1%}",
                "description": "Average improvement over baseline"
            })
        
        return pd.DataFrame(rows)
    
    def create_detailed_results_table(self) -> pd.DataFrame:
        """Create detailed results with all individual runs"""
        detailed_data = []
        
        for strategy in self.strategies:
            for batch_size in self.batch_sizes:
                file_path = self.results_dir / f"strategy_{strategy}_batch_{batch_size}.pkl"
                
                if file_path.exists():
                    try:
                        with open(file_path, 'rb') as f:
                            results = pickle.load(f)
                        
                        for result in results:
                            detailed_data.append({
                                "strategy": result.strategy,
                                "strategy_name": self.strategy_names[strategy],
                                "batch_size": result.batch_size,
                                "run_id": result.run_id,
                                "inference_latency_ms": round(result.inference_latency_ms, 3),
                                "memory_usage_mb": round(result.memory_usage_mb, 1),
                                "cache_hit_rate": round(result.cache_hit_rate, 4),
                                "expert_transfer_time_ms": round(result.expert_transfer_time_ms, 3),
                                "gpu_utilization": round(result.gpu_utilization, 3),
                                "prefetch_accuracy": round(result.prefetch_accuracy, 4),
                                "total_experts_loaded": result.total_experts_loaded,
                                "cache_misses": result.cache_misses,
                                "timestamp": result.timestamp
                            })
                    
                    except Exception as e:
                        print(f"❌ Error loading {file_path}: {e}")
        
        return pd.DataFrame(detailed_data)
    
    def save_csv_report(self, output_file: Path, analysis_data: dict):
        """Save comprehensive CSV report"""
        # Create summary table
        summary_df = self.create_summary_table(analysis_data)
        
        # Save to CSV
        summary_df.to_csv(output_file, index=False)
        print(f"📊 CSV report saved to: {output_file}")
        
        # Also save comparison matrix
        comparison_df = self.create_comparison_matrix(analysis_data)
        comparison_file = output_file.with_name(output_file.stem + "_comparison.csv")
        comparison_df.to_csv(comparison_file)
        print(f"📊 Comparison matrix saved to: {comparison_file}")
        
        # Save statistical tests
        stat_tests_df = self.create_statistical_tests_table(analysis_data)
        stat_file = output_file.with_name(output_file.stem + "_statistical_tests.csv")
        stat_tests_df.to_csv(stat_file, index=False)
        print(f"📊 Statistical tests saved to: {stat_file}")
        
        # Save detailed results
        detailed_df = self.create_detailed_results_table()
        detailed_file = output_file.with_name(output_file.stem + "_detailed.csv")
        detailed_df.to_csv(detailed_file, index=False)
        print(f"📊 Detailed results saved to: {detailed_file}")
    
    def create_excel_report(self, output_file: Path, analysis_data: dict):
        """Create formatted Excel report"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Sheet 1: Summary
        ws1 = wb.create_sheet("Summary")
        summary_df = self.create_summary_table(analysis_data)
        
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws1.append(r)
        
        # Format header
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Format data cells
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
            for cell in row:
                cell.border = border
                if cell.column in [3, 4, 5, 6]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Comparison Matrix
        ws2 = wb.create_sheet("Comparison_Matrix")
        comparison_df = self.create_comparison_matrix(analysis_data)
        
        # Add index as first column
        comparison_df.reset_index(inplace=True)
        
        for r in dataframe_to_rows(comparison_df, index=False, header=True):
            ws2.append(r)
        
        # Format header
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Statistical Tests
        ws3 = wb.create_sheet("Statistical_Tests")
        stat_df = self.create_statistical_tests_table(analysis_data)
        
        if not stat_df.empty:
            for r in dataframe_to_rows(stat_df, index=False, header=True):
                ws3.append(r)
            
            # Format header
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
        
        # Sheet 4: Key Insights
        ws4 = wb.create_sheet("Key_Insights")
        insights_df = self.create_insights_summary(analysis_data)
        
        if not insights_df.empty:
            for r in dataframe_to_rows(insights_df, index=False, header=True):
                ws4.append(r)
            
            # Format header
            for cell in ws4[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws4.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws4.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 5: Detailed Results (first 1000 rows to avoid file size issues)
        ws5 = wb.create_sheet("Detailed_Results")
        detailed_df = self.create_detailed_results_table()
        
        if not detailed_df.empty:
            # Limit to first 1000 rows for Excel performance
            detailed_subset = detailed_df.head(1000)
            
            for r in dataframe_to_rows(detailed_subset, index=False, header=True):
                ws5.append(r)
            
            # Format header
            for cell in ws5[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
        
        # Save workbook
        wb.save(output_file)
        print(f"📊 Excel report saved to: {output_file}")
    
    def generate_readme(self, output_dir: Path, analysis_data: dict):
        """Generate README with key findings"""
        readme_content = f"""# Switch Transformer Prefetching Evaluation Results

## Overview

This directory contains the results of comprehensive Switch Transformer prefetching evaluation across 5 strategies and 5 batch sizes (25 total experiments).

## Files

- `switch_results.csv` - Main summary statistics
- `switch_results_comparison.csv` - Strategy comparison matrix
- `switch_results_statistical_tests.csv` - Statistical significance tests
- `switch_results_detailed.csv` - All individual run results
- `switch_results.xlsx` - Formatted Excel report with all sheets
- `analysis.json` - Complete analysis results in JSON format

## Key Findings

"""
        
        insights = analysis_data.get("insights", {})
        if insights.get("recommendations"):
            readme_content += "### Recommendations\n\n"
            for rec in insights["recommendations"]:
                readme_content += f"- {rec}\n"
        
        readme_content += f"""

### Strategy Descriptions

- **A (On-Demand)**: Baseline - load experts from CPU memory on demand
- **B (Oracle)**: Perfect one-layer-ahead prefetching (100% accuracy)
- **C (Multi-Look)**: Multi-layer-ahead prediction with 47.55% accuracy
- **D (Top-K)**: Top-10 expert loading per layer with prediction
- **E (Intelligent)**: Strategy D + adaptive intelligent caching

### Experimental Setup

- **Batch Sizes**: 1, 2, 4, 8, 16
- **Runs per Experiment**: 10 (for statistical significance)
- **Model**: Switch Transformer Base (128 experts, 12 layers)
- **Hardware**: Calibrated for RTX 3090 timing

### Usage

Load the results in Python:
```python
import pandas as pd

# Main results
df = pd.read_csv('switch_results.csv')

# Detailed analysis
import json
with open('analysis.json') as f:
    analysis = json.load(f)
```

Or open `switch_results.xlsx` in Excel for formatted analysis.
"""
        
        readme_file = output_dir / "README.md"
        with open(readme_file, 'w') as f:
            f.write(readme_content)
        
        print(f"📄 README saved to: {readme_file}")

def main():
    parser = argparse.ArgumentParser(description="Generate CSV/Excel Report")
    parser.add_argument("--input", type=str, required=True,
                       help="Input directory with analysis results")
    parser.add_argument("--output", type=str, default=None,
                       help="Output file prefix (default: input_dir/switch_results)")
    parser.add_argument("--format", choices=["csv", "excel", "both"], default="both",
                       help="Output format")
    
    args = parser.parse_args()
    
    # Set up paths
    input_dir = Path(args.input)
    if args.output:
        output_prefix = Path(args.output)
    else:
        output_prefix = input_dir / "switch_results"
    
    # Initialize generator
    generator = ReportGenerator(input_dir)
    
    # Load analysis data
    analysis_data = generator.load_analysis_results()
    if not analysis_data:
        return
    
    # Generate reports
    if args.format in ["csv", "both"]:
        csv_file = output_prefix.with_suffix('.csv')
        generator.save_csv_report(csv_file, analysis_data)
    
    if args.format in ["excel", "both"]:
        excel_file = output_prefix.with_suffix('.xlsx')
        generator.create_excel_report(excel_file, analysis_data)
    
    # Generate README
    generator.generate_readme(input_dir, analysis_data)
    
    print(f"\n✅ Report generation completed!")
    print(f"📁 Output files in: {input_dir}")

if __name__ == "__main__":
    main()